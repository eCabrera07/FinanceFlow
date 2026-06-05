"""
One-time script — generates backend/assets/default_template.xlsx.
Run: python scripts/create_template.py
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, DoughnutChart, Reference

CATEGORIES = [
    "Food & Dining", "Gas & Fuel", "Groceries", "Shopping",
    "Subscriptions", "Entertainment", "Healthcare", "Utilities",
    "Travel", "Other",
]

MONTHS = [
    "Jan 2026", "Feb 2026", "Mar 2026", "Apr 2026",
    "May 2026", "Jun 2026", "Jul 2026", "Aug 2026",
    "Sep 2026", "Oct 2026", "Nov 2026", "Dec 2026",
]

HEADER_FONT = Font(bold=True)
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREEN  = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
RED    = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_budgets_tab(wb):
    ws = wb.create_sheet("📋 Budgets")
    ws["A1"], ws["B1"] = "Category", "Monthly Budget"
    ws["A1"].font = ws["B1"].font = HEADER_FONT
    for i, cat in enumerate(CATEGORIES, start=2):
        ws[f"A{i}"] = cat
        ws[f"B{i}"] = 0
    _set_col_widths(ws, {"A": 20, "B": 15})


def build_monthly_tab(wb, month_name: str):
    ws = wb.create_sheet(month_name)

    for col, label in enumerate(["Date", "Description", "Amount", "Category", "Source", "Type"], start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = label
        cell.font = HEADER_FONT

    ws["H1"] = "Summary"
    ws["H1"].font = HEADER_FONT
    ws["H2"] = "Total Income"
    ws["I2"] = '=SUMIF(F:F,"Income",C:C)'
    ws["H3"] = "Total Expenses"
    ws["I3"] = '=SUMIF(F:F,"Expense",C:C)'
    ws["H4"] = "Net Balance"
    ws["I4"] = "=I2+I3"
    ws["H5"] = "Uncategorized"
    ws["I5"] = '=COUNTIF(D:D,"Uncategorized")'
    ws.conditional_formatting.add("I5", CellIsRule(operator="greaterThan", formula=["0"], fill=YELLOW))

    ws["H7"] = "Category"
    ws["I7"] = "Subtotal"
    ws["H7"].font = ws["I7"].font = HEADER_FONT
    for i, cat in enumerate(CATEGORIES, start=8):
        ws[f"H{i}"] = cat
        ws[f"I{i}"] = f'=SUMIF(D:D,"{cat}",C:C)'

    _set_col_widths(ws, {"A": 12, "B": 32, "C": 12, "D": 20, "E": 20, "F": 10, "H": 20, "I": 14})
    return ws


def build_dashboard_tab(wb):
    ws = wb.create_sheet("📊 Dashboard")

    ws["A1"] = "Year Summary"
    ws["A1"].font = Font(bold=True, size=13)
    for col, label in zip("ABCD", ["Month", "Expenses", "Income", "Net"]):
        cell = ws[f"{col}2"]
        cell.value = label
        cell.font = HEADER_FONT

    for i, month in enumerate(MONTHS, start=3):
        ws[f"A{i}"] = month
        ws[f"B{i}"] = f'=IFERROR(INDIRECT("\'"&A{i}&"\'!I3"),0)'
        ws[f"C{i}"] = f'=IFERROR(INDIRECT("\'"&A{i}&"\'!I2"),0)'
        ws[f"D{i}"] = f"=C{i}+B{i}"

    ws["F1"] = "Budget Progress"
    ws["F1"].font = Font(bold=True, size=13)
    ws["K1"] = "Viewing month:"
    ws["L1"] = "Jan 2026"
    ws["L1"].font = Font(italic=True)

    for col, label in zip("FGHI", ["Category", "Spent", "Budget", "% Used"]):
        cell = ws[f"{col}2"]
        cell.value = label
        cell.font = HEADER_FONT

    for i, cat in enumerate(CATEGORIES, start=3):
        ws[f"F{i}"] = cat
        ws[f"G{i}"] = (
            f'=IFERROR(SUMIF(INDIRECT("\'"&$L$1&"\'!H:H"),F{i},'
            f'INDIRECT("\'"&$L$1&"\'!I:I")),0)'
        )
        ws[f"H{i}"] = f'=IFERROR(VLOOKUP(F{i},\'📋 Budgets\'!A:B,2,0),0)'
        ws[f"I{i}"] = f'=IFERROR(ABS(G{i})/H{i},0)'

    budget_range = f"I3:I{2 + len(CATEGORIES)}"
    ws.conditional_formatting.add(budget_range, CellIsRule(operator="lessThanOrEqual", formula=["0.8"], fill=GREEN))
    ws.conditional_formatting.add(budget_range, CellIsRule(operator="greaterThan",     formula=["1.0"], fill=RED))

    bar = BarChart()
    bar.type = "col"
    bar.title = "Monthly Spending"
    bar.y_axis.title = "Amount ($)"
    bar.add_data(Reference(ws, min_col=2, min_row=2, max_row=14), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=3, max_row=14))
    bar.width, bar.height = 18, 12
    ws.add_chart(bar, "A16")

    grouped = BarChart()
    grouped.type = "col"
    grouped.grouping = "clustered"
    grouped.title = "Income vs Expenses"
    grouped.y_axis.title = "Amount ($)"
    grouped.add_data(Reference(ws, min_col=2, min_row=2, max_col=3, max_row=14), titles_from_data=True)
    grouped.set_categories(Reference(ws, min_col=1, min_row=3, max_row=14))
    grouped.width, grouped.height = 18, 12
    ws.add_chart(grouped, "J16")

    donut = DoughnutChart()
    donut.title = "Spending by Category"
    donut.add_data(Reference(ws, min_col=7, min_row=3, max_row=2 + len(CATEGORIES)))
    donut.set_categories(Reference(ws, min_col=6, min_row=3, max_row=2 + len(CATEGORIES)))
    donut.width, donut.height = 14, 14
    ws.add_chart(donut, "A33")

    _set_col_widths(ws, {"A": 12, "B": 14, "C": 14, "D": 10, "F": 20, "G": 14, "H": 12, "I": 10, "K": 16, "L": 12})
    return ws


def main():
    wb = openpyxl.Workbook()
    del wb["Sheet"]  # remove default blank sheet

    build_budgets_tab(wb)
    build_dashboard_tab(wb)

    # Hidden _template tab — writer copies this every time a new month tab is needed.
    build_monthly_tab(wb, "_template")
    wb["_template"].sheet_state = "hidden"

    # Starter visible month tab — created by copying _template so it's identical.
    jan = wb.copy_worksheet(wb["_template"])
    jan.title = "Jan 2026"

    out = os.path.join(os.path.dirname(__file__), "..", "assets", "default_template.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"Saved: {os.path.abspath(out)}")


if __name__ == "__main__":
    main()
