import openpyxl
from typing import Dict


def read_spreadsheet_structure(file_path: str) -> Dict[str, Dict[str, str]]:
    """Read sheet names and first-row headers from an xlsx file.

    Returns:
        {sheet_name: {column_letter: header_text}}
        Only columns with non-empty headers are included.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        result = {}
        for name in wb.sheetnames:
            ws = wb[name]
            headers = {}
            for cell in ws[1]:
                if cell.value is not None:
                    headers[cell.column_letter] = str(cell.value)
            result[name] = headers
        return result
    finally:
        wb.close()
