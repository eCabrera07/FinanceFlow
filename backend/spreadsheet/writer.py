import openpyxl
from openpyxl.utils import column_index_from_string
from typing import Any, Dict, List, Optional

_DEFAULT_COLUMNS = {
    "date": "A",
    "description": "B",
    "amount": "C",
    "category": "D",
    "source": "E",
    "type": "F",
}


def write_transactions(
    file_path: str,
    sheet_name: str,
    transactions: List[Dict[str, Any]],
    mapping: Optional[Dict[str, Any]] = None,
) -> None:
    """Write transaction rows to an xlsx file, skipping duplicates.

    A transaction is a duplicate if an existing row shares the same
    date, description, and amount. Duplicates are silently skipped.
    """
    wb = openpyxl.load_workbook(file_path)

    target_sheet = mapping["sheet_name"] if mapping else sheet_name
    if target_sheet not in wb.sheetnames:
        if "_template" in wb.sheetnames:
            new_ws = wb.copy_worksheet(wb["_template"])
            new_ws.title = target_sheet
        else:
            wb.create_sheet(target_sheet)
    ws = wb[target_sheet]

    col_map = mapping["columns"] if mapping else _DEFAULT_COLUMNS

    if mapping and str(mapping.get("start_row", "auto")) != "auto":
        start_row = int(mapping["start_row"])
    else:
        start_row = _first_empty_row_in_col_a(ws)

    existing = _existing_keys(ws, col_map)
    write_row = start_row

    for tx in transactions:
        key = _key(tx)
        if key in existing:
            continue
        for field, col_letter in col_map.items():
            if col_letter is None:
                continue
            ws.cell(row=write_row, column=column_index_from_string(col_letter), value=tx.get(field))
        existing.add(key)
        write_row += 1

    wb.save(file_path)


def _key(tx: Dict[str, Any]) -> tuple:
    return (str(tx.get("date", "")), str(tx.get("description", "")), float(tx.get("amount", 0)))


def _existing_keys(ws, col_map: Dict[str, Any]) -> set:
    """Read all existing rows and return a set of (date, description, amount) tuples."""
    date_col = col_map.get("date") or "A"
    desc_col = col_map.get("description") or "B"
    amt_col = col_map.get("amount") or "C"
    keys = set()
    for row in ws.iter_rows(min_row=1):
        d = ws[f"{date_col}{row[0].row}"].value
        desc = ws[f"{desc_col}{row[0].row}"].value
        amt = ws[f"{amt_col}{row[0].row}"].value
        if d is not None and desc is not None and amt is not None:
            try:
                keys.add((str(d), str(desc), float(amt)))
            except (ValueError, TypeError):
                continue
    return keys


def _first_empty_row_in_col_a(ws) -> int:
    for (cell,) in ws.iter_rows(min_col=1, max_col=1):
        if cell.value is None:
            return cell.row
    max_row = ws.max_row
    return (max_row + 1) if max_row is not None else 1
