import openpyxl
from spreadsheet.writer import write_transactions


def base_xlsx(tmp_path, sheet="Jan 2026") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    path = str(tmp_path / "test.xlsx")
    wb.save(path)
    return path


def make_tx(**overrides):
    return {"date": "2026-01-15", "description": "Netflix", "amount": -15.99,
            "category": "Subscriptions", "source": "Chase", "type": "Expense", **overrides}


def test_duplicate_not_written_twice(tmp_path):
    path = base_xlsx(tmp_path)
    tx = make_tx()
    write_transactions(path, "Jan 2026", [tx], mapping=None)
    write_transactions(path, "Jan 2026", [tx], mapping=None)

    wb = openpyxl.load_workbook(path)
    ws = wb["Jan 2026"]
    assert ws["A1"].value == "2026-01-15"
    assert ws["A2"].value is None  # second write was skipped


def test_different_descriptions_both_written(tmp_path):
    path = base_xlsx(tmp_path)
    write_transactions(path, "Jan 2026", [make_tx(description="Netflix")], mapping=None)
    write_transactions(path, "Jan 2026", [make_tx(description="Spotify")], mapping=None)

    wb = openpyxl.load_workbook(path)
    ws = wb["Jan 2026"]
    assert ws["B1"].value == "Netflix"
    assert ws["B2"].value == "Spotify"


def test_same_desc_different_amount_both_written(tmp_path):
    path = base_xlsx(tmp_path)
    write_transactions(path, "Jan 2026", [make_tx(amount=-15.99)], mapping=None)
    write_transactions(path, "Jan 2026", [make_tx(amount=-16.99)], mapping=None)

    wb = openpyxl.load_workbook(path)
    ws = wb["Jan 2026"]
    assert ws["C1"].value == -15.99
    assert ws["C2"].value == -16.99
