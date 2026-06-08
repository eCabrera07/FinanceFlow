import io
import json
import os
import openpyxl
import pytest
from fastapi.testclient import TestClient
from main import app

client = TestClient(app)


@pytest.fixture(autouse=True)
def isolated_volume(tmp_path, monkeypatch):
    """Redirect VOLUME_XLSX_PATH to a tmp dir so tests never touch real data."""
    import statement.router as statement_router
    monkeypatch.setattr(
        statement_router,
        "VOLUME_XLSX_PATH",
        str(tmp_path / "FinanceFlow.xlsx"),
    )


CHASE_CSV = """Transaction Date,Post Date,Description,Category,Type,Amount,Memo
01/15/2026,01/16/2026,NETFLIX.COM,Entertainment,Sale,-15.99,
01/16/2026,01/17/2026,DIRECT DEPOSIT,Income,Payment,2500.00,
"""

SAMPLE_MAPPING = {
    "file_path": "finances.xlsx",
    "sheet_name": "Transactions",
    "start_row": "auto",
    "columns": {"date": "A", "description": "B", "amount": "C",
                "category": "D", "source": "E", "type": "F"},
}


def xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    wb.active.title = "Transactions"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_upload_csv_returns_transactions():
    res = client.post(
        "/statement/upload",
        files={"file": ("chase.csv", io.BytesIO(CHASE_CSV.encode()), "text/csv")},
    )
    assert res.status_code == 200
    data = res.json()
    assert len(data["transactions"]) == 2


def test_upload_transaction_has_required_fields():
    res = client.post(
        "/statement/upload",
        files={"file": ("chase.csv", io.BytesIO(CHASE_CSV.encode()), "text/csv")},
    )
    tx = res.json()["transactions"][0]
    for field in ("date", "description", "amount", "category", "source", "type"):
        assert field in tx


def test_upload_unsupported_format_returns_400():
    res = client.post(
        "/statement/upload",
        files={"file": ("notes.txt", io.BytesIO(b"hello"), "text/plain")},
    )
    assert res.status_code == 400


def test_confirm_without_mapping_uses_default_template():
    txs = [{"date": "01/15/2026", "description": "Netflix", "amount": -15.99,
             "category": "Subscriptions", "source": "Chase", "type": "Expense"}]
    wb = openpyxl.Workbook()
    wb.active.title = "Jan 2026"
    buf = io.BytesIO()
    wb.save(buf)
    res = client.post(
        "/statement/confirm",
        data={"transactions": json.dumps(txs)},
        files={"spreadsheet": ("f.xlsx", io.BytesIO(buf.getvalue()),
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]


def test_confirm_with_mapping_returns_xlsx():
    client.post("/spreadsheet/mapping", json=SAMPLE_MAPPING)

    txs = [{"date": "2026-01-15", "description": "Netflix", "amount": -15.99,
             "category": "Subscriptions", "source": "Chase", "type": "Expense"}]
    res = client.post(
        "/statement/confirm",
        data={"transactions": json.dumps(txs)},
        files={"spreadsheet": ("f.xlsx", io.BytesIO(xlsx_bytes()),
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    assert len(res.content) > 1000

    client.delete("/spreadsheet/mapping")


def test_confirm_uses_volume_file_when_present(tmp_path, monkeypatch):
    """Case 1: volume file exists — write in-place, return JSON."""
    import statement.router as statement_router
    volume_path = str(tmp_path / "FinanceFlow.xlsx")
    monkeypatch.setattr(statement_router, "VOLUME_XLSX_PATH", volume_path)

    # Pre-create the volume xlsx
    wb = openpyxl.Workbook()
    wb.active.title = "Jun 2026"
    wb.save(volume_path)

    txs = [{"date": "06/01/2026", "description": "Groceries", "amount": -42.00,
             "category": "Groceries", "source": "Chase", "type": "Expense"}]
    res = client.post("/statement/confirm", data={"transactions": json.dumps(txs)})
    assert res.status_code == 200
    assert res.json()["status"] == "written"
    # File should still exist and contain the transaction
    wb2 = openpyxl.load_workbook(volume_path)
    ws = wb2["Jun 2026"]
    assert ws["A1"].value == "06/01/2026"


def test_confirm_saves_uploaded_file_to_volume(tmp_path, monkeypatch):
    """Case 2: no volume file, spreadsheet uploaded — save to volume, return download."""
    import statement.router as statement_router
    volume_path = str(tmp_path / "FinanceFlow.xlsx")
    monkeypatch.setattr(statement_router, "VOLUME_XLSX_PATH", volume_path)

    wb = openpyxl.Workbook()
    wb.active.title = "Jun 2026"
    buf = io.BytesIO()
    wb.save(buf)

    txs = [{"date": "06/01/2026", "description": "Netflix", "amount": -15.99,
             "category": "Subscriptions", "source": "Chase", "type": "Expense"}]
    res = client.post(
        "/statement/confirm",
        data={"transactions": json.dumps(txs)},
        files={"spreadsheet": ("f.xlsx", io.BytesIO(buf.getvalue()),
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    # Volume file should now exist for future requests
    assert os.path.exists(volume_path)


def test_confirm_creates_default_when_no_file_anywhere(tmp_path, monkeypatch):
    """Case 3: no volume file, no upload — create from template, return JSON."""
    import statement.router as statement_router
    volume_path = str(tmp_path / "FinanceFlow.xlsx")
    monkeypatch.setattr(statement_router, "VOLUME_XLSX_PATH", volume_path)

    txs = [{"date": "06/01/2026", "description": "Salary", "amount": 3000.00,
             "category": "Other", "source": "Employer", "type": "Income"}]
    res = client.post("/statement/confirm", data={"transactions": json.dumps(txs)})
    assert res.status_code == 200
    assert res.json()["status"] == "created"
    # Default template should now exist at volume path
    assert os.path.exists(volume_path)
    wb = openpyxl.load_workbook(volume_path)
    assert len(wb.sheetnames) >= 1
