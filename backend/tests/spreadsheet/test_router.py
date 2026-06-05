import pytest
from fastapi.testclient import TestClient
from main import app

client = TestClient(app)


def test_download_template_returns_xlsx():
    response = client.get("/spreadsheet/template/download")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert len(response.content) > 1000  # not an empty file


def test_download_template_returns_500_when_missing(monkeypatch):
    from spreadsheet import template_service
    monkeypatch.setattr(template_service, "TEMPLATE_PATH", "/nonexistent/path/file.xlsx")
    response = client.get("/spreadsheet/template/download")
    assert response.status_code == 500
    assert "detail" in response.json()


import io
import openpyxl


def test_import_returns_sheet_structure():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws["A1"] = "Date"
    ws["B1"] = "Amount"
    ws["C1"] = "Description"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = client.post(
        "/spreadsheet/import",
        files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    data = response.json()
    assert "Transactions" in data
    assert data["Transactions"]["headers"]["A"] == "Date"
    assert data["Transactions"]["suggested_mapping"]["date"] == "A"
    assert data["Transactions"]["suggested_mapping"]["amount"] == "B"


SAMPLE_MAPPING = {
    "file_path": "/test/finances.xlsx",
    "sheet_name": "Transactions",
    "start_row": "auto",
    "columns": {"date": "A", "description": "B", "amount": "C", "category": "D", "source": None, "type": None},
}


def test_save_mapping_returns_saved():
    response = client.post("/spreadsheet/mapping", json=SAMPLE_MAPPING)
    assert response.status_code == 200
    assert response.json()["status"] == "saved"


def test_get_mapping_returns_saved_data():
    client.post("/spreadsheet/mapping", json=SAMPLE_MAPPING)
    response = client.get("/spreadsheet/mapping")
    assert response.status_code == 200
    assert response.json()["mapping"]["sheet_name"] == "Transactions"


def test_delete_mapping_clears_it():
    client.post("/spreadsheet/mapping", json=SAMPLE_MAPPING)
    client.delete("/spreadsheet/mapping")
    response = client.get("/spreadsheet/mapping")
    assert response.json()["mapping"] is None
