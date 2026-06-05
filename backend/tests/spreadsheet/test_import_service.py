import openpyxl
import pytest
from spreadsheet.import_service import read_spreadsheet_structure


def make_xlsx(tmp_path, sheets: dict) -> str:
    """Helper: create an xlsx with given {sheet_name: {col_letter: header}}."""
    wb = openpyxl.Workbook()
    first = True
    for name, headers in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for col, text in headers.items():
            ws[f"{col}1"] = text
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return str(path)


def test_reads_single_sheet_headers(tmp_path):
    path = make_xlsx(tmp_path, {"Transactions": {"A": "Date", "B": "Amount", "C": "Payee"}})
    result = read_spreadsheet_structure(path)
    assert "Transactions" in result
    assert result["Transactions"] == {"A": "Date", "B": "Amount", "C": "Payee"}


def test_reads_multiple_sheets(tmp_path):
    path = make_xlsx(tmp_path, {"Sheet1": {"A": "Date"}, "Sheet2": {"A": "Merchant"}})
    result = read_spreadsheet_structure(path)
    assert "Sheet1" in result
    assert "Sheet2" in result


def test_ignores_empty_columns(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Date"
    ws["C1"] = "Amount"  # B1 is empty
    path = tmp_path / "sparse.xlsx"
    wb.save(str(path))
    result = read_spreadsheet_structure(str(path))
    assert "B" not in result["Data"]
    assert result["Data"]["A"] == "Date"
    assert result["Data"]["C"] == "Amount"
