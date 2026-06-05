import os
import pytest
import openpyxl
from spreadsheet.template_service import get_template_path, TEMPLATE_PATH


def test_template_file_exists():
    assert os.path.exists(TEMPLATE_PATH), f"Template not found at {TEMPLATE_PATH}"


def test_template_has_required_sheets():
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    assert "📋 Budgets" in wb.sheetnames
    assert "📊 Dashboard" in wb.sheetnames
    assert "_template" in wb.sheetnames
    assert "Jan 2026" in wb.sheetnames
    wb.close()


def test_template_tab_is_hidden():
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    assert wb["_template"].sheet_state == "hidden"
    wb.close()


def test_template_tab_has_month_structure():
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["_template"]
    assert ws["A1"].value == "Date"
    assert ws["H1"].value == "Summary"
    wb.close()


def test_get_template_path_returns_path():
    path = get_template_path()
    assert path == TEMPLATE_PATH
    assert path.endswith(".xlsx")
