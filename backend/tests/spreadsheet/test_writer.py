import openpyxl
import pytest
from spreadsheet.writer import write_transactions


def make_transaction(**overrides):
    base = {
        "date": "2026-01-15",
        "description": "Walmart",
        "amount": -52.40,
        "category": "Groceries",
        "source": "Chase",
        "type": "Expense",
    }
    return {**base, **overrides}


def empty_xlsx(tmp_path, sheet_name="Jan 2026") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    path = str(tmp_path / "test.xlsx")
    wb.save(path)
    return path


def test_writes_to_default_columns(tmp_path):
    path = empty_xlsx(tmp_path)
    write_transactions(path, "Jan 2026", [make_transaction()], mapping=None)

    wb = openpyxl.load_workbook(path)
    ws = wb["Jan 2026"]
    assert ws["A1"].value == "2026-01-15"
    assert ws["B1"].value == "Walmart"
    assert ws["C1"].value == 52.40
    assert ws["D1"].value == "Groceries"
    assert ws["E1"].value == "Chase"
    assert ws["F1"].value == "Expense"


def test_appends_below_existing_rows(tmp_path):
    path = empty_xlsx(tmp_path)
    write_transactions(path, "Jan 2026", [make_transaction(description="First")], mapping=None)
    write_transactions(path, "Jan 2026", [make_transaction(description="Second")], mapping=None)

    wb = openpyxl.load_workbook(path)
    ws = wb["Jan 2026"]
    assert ws["B1"].value == "First"
    assert ws["B2"].value == "Second"


def test_writes_to_mapped_columns(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MySheet"
    ws["C1"] = "Date Header"  # existing header row
    path = str(tmp_path / "custom.xlsx")
    wb.save(path)

    mapping = {
        "sheet_name": "MySheet",
        "start_row": "3",
        "columns": {
            "date": "C",
            "description": "D",
            "amount": "E",
            "category": None,
            "source": None,
            "type": None,
        },
    }
    write_transactions(path, "MySheet", [make_transaction()], mapping=mapping)

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["MySheet"]
    assert ws2["C3"].value == "2026-01-15"
    assert ws2["D3"].value == "Walmart"
    assert ws2["E3"].value == 52.40
    assert ws2["C1"].value == "Date Header"  # existing data untouched


def test_skips_none_columns_in_mapping(tmp_path):
    path = empty_xlsx(tmp_path, "Data")
    mapping = {
        "sheet_name": "Data",
        "start_row": "1",
        "columns": {"date": "A", "description": None, "amount": "B", "category": None, "source": None, "type": None},
    }
    write_transactions(path, "Data", [make_transaction()], mapping=mapping)

    wb = openpyxl.load_workbook(path)
    ws = wb["Data"]
    assert ws["A1"].value == "2026-01-15"
    assert ws["B1"].value == 52.40
    assert ws["C1"].value is None  # no column C written


def test_creates_sheet_if_missing(tmp_path):
    wb = openpyxl.Workbook()
    path = str(tmp_path / "blank.xlsx")
    wb.save(path)

    write_transactions(path, "Feb 2026", [make_transaction()], mapping=None)

    wb2 = openpyxl.load_workbook(path)
    assert "Feb 2026" in wb2.sheetnames


def test_new_month_tab_copied_from_template(tmp_path):
    """When the xlsx has a _template tab, new month tabs should be copied from it
    so they inherit the header row and H:I summary formulas."""
    import shutil
    from spreadsheet.template_service import TEMPLATE_PATH

    dest = str(tmp_path / "finances.xlsx")
    shutil.copy2(TEMPLATE_PATH, dest)

    write_transactions(dest, "Feb 2026", [make_transaction()], mapping=None)

    wb = openpyxl.load_workbook(dest)
    assert "Feb 2026" in wb.sheetnames
    ws = wb["Feb 2026"]
    # Header row inherited from _template
    assert ws["A1"].value == "Date"
    assert ws["H1"].value == "📊 Monthly Summary"
    # Transaction written at row 2 (row 1 is the header)
    assert ws["A2"].value == "2026-01-15"
    assert ws["B2"].value == "Walmart"
